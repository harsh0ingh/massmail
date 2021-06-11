import smtplib
import openpyxl
import sys
from pprint import pprint
from email import encoders
from email.message import EmailMessage


wb = openpyxl.load_workbook('mail_list.xlsx')
sheet = wb['Sheet1']


sender_mail = ""        #sender's mail goes here
password = ""           #password goes here
carbon_copy_mail = ""   #cc mail
subject = ""            #subject
messsage = ""           #body


def send_email(email_address, k):
    try:

        print(email_address)

        smtpObj = smtplib.SMTP('smtp.gmail.com', '587')
        smtpObj.ehlo()
        smtpObj.starttls()
        smtpObj.login(mymail, password)
        print("Sending mail to ", email_address)

        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = mymail
        msg['To'] = email_address
        msg['Cc'] = ccmail
       
        msg.set_content(messsage) 

        smtpObj.send_message(msg)
        print("successful")
        
    except:
            pprint("failed")

for i in range(2, sheet.max_row+2):
    
        send_email(sheet.cell(row=i, column=2).value, i)



# Attaching pdf file
    #       file_data1 = f1.read()
    #       file_name1 = f1.name
    #       msg.add_attachment(file_data1, maintype = 'application', subtype = 'octet-stream', filename = file_name1)
    
# Attaching image file
    #       f2 = open("filename.img", 'rb')    
    #       file_data2 = f2.read()
    #       file_name2 = f2.name
    #       msg.add_attachment(file_data2, maintype = 'image', subtype = 'octet-stream', filename = file_name2) 

 

# Attaching html page and links
    #   msg.add_alternative("""\
    #   Dear Madam/Sir, <br>
    #   This is a test mail. Pls ignore.<br>
    #   Thanking you,<br>
    #   Sincerely,<br>
    #   Test Mail.<br>

    #   <!DOCTYPE html>
    #   <html>
    #    <body>
    #        <a href="https://drive.google.com/file/d/1KW41PpKVeOYIqHH3shS9_URn44L63Md4/view?usp=sharing">Expansion of Sin(X)</a>
    #        <br>
    #        <a href="https://drive.google.com/file/d/1avcwb5VWMYhsrJ2Oxr-wtvjVMl3iu0Oe/view?usp=sharing">Expansion of Sin(X)</a>
    #        <br>
    #    </body>
    #   </html>            

    #           """, subtype = 'html')